import os
import tempfile
import unittest
from unittest.mock import patch
from pathlib import Path

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
_TEST_LOCALAPPDATA = tempfile.mkdtemp(prefix="timertaskmaster-ui-")
os.environ["LOCALAPPDATA"] = _TEST_LOCALAPPDATA

from PySide6.QtCore import QDate
from PySide6.QtWidgets import QApplication, QMessageBox

import app as combined_app
from app import ExportOptionsDialog, MainWindow
from database import Database
from master_database import MasterDatabase
from timer_app import app_data_dir
from csv_store import append_audit_action, append_record


class UiSmokeTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.qt_app = QApplication.instance() or QApplication([])

    def test_export_options_dialog_opens_without_checkbox_wordwrap_error(self):
        dialog = ExportOptionsDialog(
            None,
            __import__("datetime").date(2026, 7, 13),
            {
                "projeto": "Todos",
                "tipo": "Todos",
                "origem": "Todos",
                "status": "Todos",
            },
        )
        self.assertIn("origem e status", dialog.apply_filters_check.text())
        self.assertTrue(dialog.apply_filters_check.toolTip())
        dialog.close()

    def test_combined_window_keeps_timer_and_adds_management(self):
        db_path = app_data_dir() / "timertask.db"
        timer_db = Database(db_path)
        timer_db.set_setting("user_name", "Usuário Teste")
        timer_db.set_setting("base_folder", str(Path(_TEST_LOCALAPPDATA) / "registros"))
        master_db = MasterDatabase(db_path)

        window = MainWindow(timer_db, master_db)
        tabs = [window.tabs.tabText(index) for index in range(window.tabs.count())]
        self.assertEqual(
            tabs,
            [
                "Dashboard",
                "Timer",
                "Registro manual",
                "Histórico",
                "Cadastros",
                "Usuários monitorados",
                "Configurações",
            ],
        )
        self.assertEqual(len(master_db.list_users()), 1)
        window.force_quit = True
        window.close()


    def test_deleted_record_is_red_and_not_counted_in_dashboard_or_history(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            db_path = root / "timertask.db"
            csv_base = root / "base"
            timer_db = Database(db_path)
            timer_db.set_setting("user_name", "Usuário Teste")
            timer_db.set_setting("base_folder", str(csv_base))
            record = {
                "registro_id": "deleted-ui-1",
                "usuario": "Usuário Teste",
                "origem_registro": "TIMER",
                "projeto": "Projeto UI",
                "tipo_atividade": "Teste",
                "descricao": "Registro incorreto",
                "inicio": "2026-07-13 08:00:00",
                "fim": "2026-07-13 09:00:00",
                "duracao_segundos": 3600,
                "duracao_formatada": "01:00:00",
                "observacao": "",
                "computador": "PC",
                "data_registro": "2026-07-13 09:00:00",
            }
            action = {
                "acao_id": "deleted-ui-action-1",
                "registro_id": "deleted-ui-1",
                "acao": "EXCLUIR",
                "data_hora_acao": "2026-07-13 10:00:00",
                "usuario_acao": "Usuário Teste",
                "motivo": "Timer iniciado por engano",
                "computador": "PC",
                "projeto": "Projeto UI",
                "tipo_atividade": "Teste",
                "descricao": "Registro incorreto",
                "inicio": "2026-07-13 08:00:00",
                "fim": "2026-07-13 09:00:00",
                "duracao_segundos": 3600,
                "duracao_formatada": "01:00:00",
                "origem_registro": "TIMER",
                "observacao": "",
                "data_registro": "2026-07-13 09:00:00",
            }
            append_record(str(csv_base), record)
            append_audit_action(str(csv_base), action)
            master_db = MasterDatabase(db_path)
            window = MainWindow(timer_db, master_db)
            window.date_edit.setDate(QDate(2026, 7, 13))
            window.refresh_dashboard()
            self.assertEqual(window.total_hours_label.text(), "00:00:00")
            self.assertEqual(window.records_label.text(), "0")
            self.assertEqual(window.deleted_label.text(), "1")
            user_item = window.dashboard_tree.topLevelItem(0)
            project_item = user_item.child(0)
            record_item = project_item.child(0)
            self.assertEqual(record_item.text(4), "EXCLUÍDO")

            window.history_date.setDate(QDate(2026, 7, 13))
            window.history_status_filter.setCurrentText("Excluídos")
            window.refresh_history()
            self.assertEqual(window.history_table.rowCount(), 1)
            self.assertEqual(window.history_table.item(0, 7).text(), "EXCLUÍDO")
            self.assertEqual(window.history_total_label.text(), "Total válido: 00:00:00")
            window.force_quit = True
            window.close()


    def test_delete_action_preserves_original_and_creates_audit(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            db_path = root / "timertask.db"
            csv_base = root / "base"
            timer_db = Database(db_path)
            timer_db.set_setting("user_name", "Usuário Teste")
            timer_db.set_setting("base_folder", str(csv_base))
            record = {
                "registro_id": "delete-flow-1",
                "usuario": "Usuário Teste",
                "origem_registro": "MANUAL",
                "projeto": "Projeto Fluxo",
                "tipo_atividade": "Documentação",
                "descricao": "Lançamento incorreto",
                "inicio": "2026-07-13 13:00:00",
                "fim": "2026-07-13 14:00:00",
                "duracao_segundos": 3600,
                "duracao_formatada": "01:00:00",
                "observacao": "",
                "computador": "PC",
                "data_registro": "2026-07-13 14:00:00",
            }
            original_path = append_record(str(csv_base), record)
            master_db = MasterDatabase(db_path)
            window = MainWindow(timer_db, master_db)
            window.history_date.setDate(QDate(2026, 7, 13))
            window.refresh_history()
            self.assertEqual(window.history_table.rowCount(), 1)
            window.history_table.selectRow(0)

            with (
                patch("timer_app.QInputDialog.getMultiLineText", return_value=("Registro duplicado", True)),
                patch("timer_app.QMessageBox.question", return_value=QMessageBox.StandardButton.Yes),
                patch("timer_app.QMessageBox.information"),
            ):
                window.delete_selected_history_record()

            self.assertTrue(original_path.exists())
            import csv
            with original_path.open("r", newline="", encoding="utf-8-sig") as handle:
                self.assertEqual(len(list(csv.DictReader(handle, delimiter=";"))), 1)
            actions = timer_db.list_audit_actions()
            self.assertEqual(len(actions), 1)
            self.assertEqual(actions[0]["status"], "SINCRONIZADO")
            self.assertEqual(actions[0]["data"]["motivo"], "Registro duplicado")
            window.history_status_filter.setCurrentText("Excluídos")
            window.refresh_history()
            self.assertEqual(window.history_table.rowCount(), 1)
            self.assertEqual(window.history_total_label.text(), "Total válido: 00:00:00")
            window.force_quit = True
            window.close()

    def test_dashboard_export_button_creates_filtered_excel(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            db_path = root / "timertask.db"
            csv_base = root / "base"
            target = root / "relatorio-dashboard.xlsx"
            timer_db = Database(db_path)
            timer_db.set_setting("user_name", "Usuário Teste")
            timer_db.set_setting("base_folder", str(csv_base))
            append_record(
                str(csv_base),
                {
                    "registro_id": "export-ui-1",
                    "usuario": "Usuário Teste",
                    "origem_registro": "TIMER",
                    "projeto": "Projeto Exportação",
                    "tipo_atividade": "Teste",
                    "descricao": "Validar botão exportar",
                    "inicio": "2026-07-13 08:00:00",
                    "fim": "2026-07-13 09:00:00",
                    "duracao_segundos": 3600,
                    "duracao_formatada": "01:00:00",
                    "observacao": "OK",
                    "computador": "PC",
                    "data_registro": "2026-07-13 09:00:00",
                },
            )
            append_record(
                str(csv_base),
                {
                    "registro_id": "export-ui-2",
                    "usuario": "Usuário Teste",
                    "origem_registro": "TIMER",
                    "projeto": "Outro Projeto",
                    "tipo_atividade": "Atendimento",
                    "descricao": "Não deve entrar no filtro",
                    "inicio": "2026-07-13 10:00:00",
                    "fim": "2026-07-13 10:30:00",
                    "duracao_segundos": 1800,
                    "duracao_formatada": "00:30:00",
                    "observacao": "",
                    "computador": "PC",
                    "data_registro": "2026-07-13 10:30:00",
                },
            )
            master_db = MasterDatabase(db_path)
            window = MainWindow(timer_db, master_db)
            self.assertEqual(window.export_button.text(), "Exportar")
            window.date_edit.setDate(QDate(2026, 7, 13))
            window.refresh_dashboard()
            window.project_filter.setCurrentText("Projeto Exportação")

            with (
                patch.object(
                    window,
                    "_get_export_options",
                    return_value={
                        "mode": "dashboard",
                        "year": 2026,
                        "months": [7],
                        "apply_dashboard_filters": True,
                        "format": "xlsx",
                    },
                ),
                patch(
                    "app.QFileDialog.getSaveFileName",
                    return_value=(str(target), "Excel completo (*.xlsx)"),
                ),
                patch("app.QMessageBox.information"),
            ):
                window.export_dashboard()

            self.assertTrue(target.exists())
            from openpyxl import load_workbook
            workbook = load_workbook(target, read_only=True)
            self.assertIn("Resumo", workbook.sheetnames)
            self.assertEqual(workbook["Registros"].max_row, 2)
            self.assertEqual(workbook["Registros"]["C2"].value, "Projeto Exportação")
            workbook.close()
            window.force_quit = True
            window.close()

    def test_dashboard_export_can_select_multiple_months(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            db_path = root / "timertask.db"
            csv_base = root / "base"
            target = root / "relatorio-semestral.xlsx"
            timer_db = Database(db_path)
            timer_db.set_setting("user_name", "Usuário Teste")
            timer_db.set_setting("base_folder", str(csv_base))
            for record_id, month in (("period-1", 1), ("period-2", 6), ("period-3", 7)):
                append_record(
                    str(csv_base),
                    {
                        "registro_id": record_id,
                        "usuario": "Usuário Teste",
                        "origem_registro": "TIMER",
                        "projeto": "Projeto Período",
                        "tipo_atividade": "Teste",
                        "descricao": "Registro mensal",
                        "inicio": f"2026-{month:02d}-13 08:00:00",
                        "fim": f"2026-{month:02d}-13 09:00:00",
                        "duracao_segundos": 3600,
                        "duracao_formatada": "01:00:00",
                        "observacao": "",
                        "computador": "PC",
                        "data_registro": f"2026-{month:02d}-13 09:00:00",
                    },
                )
            master_db = MasterDatabase(db_path)
            window = MainWindow(timer_db, master_db)

            with (
                patch.object(
                    window,
                    "_get_export_options",
                    return_value={
                        "mode": "period",
                        "year": 2026,
                        "months": [1, 2, 3, 4, 5, 6],
                        "apply_dashboard_filters": False,
                        "format": "xlsx",
                    },
                ),
                patch(
                    "app.QFileDialog.getSaveFileName",
                    return_value=(str(target), "Excel completo (*.xlsx)"),
                ),
                patch("app.QMessageBox.information"),
            ):
                window.export_dashboard()

            from openpyxl import load_workbook
            workbook = load_workbook(target, read_only=True)
            self.assertEqual(workbook["Registros"].max_row, 3)
            self.assertEqual(workbook["Resumo"]["B4"].value, "Janeiro a Junho de 2026")
            workbook.close()
            window.force_quit = True
            window.close()

    def test_upgrade_backup_copies_existing_database(self):
        source_dir = Path(tempfile.mkdtemp(prefix="timertaskmaster-backup-"))
        db_path = source_dir / "timertask.db"
        db_path.write_bytes(b"database-test")
        original_app_data_dir = combined_app.app_data_dir
        try:
            combined_app.app_data_dir = lambda: source_dir
            backup_dir = combined_app.backup_existing_timer_data()
            self.assertIsNotNone(backup_dir)
            self.assertEqual((backup_dir / "timertask.db").read_bytes(), b"database-test")
            self.assertIsNone(combined_app.backup_existing_timer_data())
        finally:
            combined_app.app_data_dir = original_app_data_dir


if __name__ == "__main__":
    unittest.main()
