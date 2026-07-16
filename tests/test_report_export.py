import csv
import tempfile
import unittest
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.chart.axis import DateAxis

from csv_reader import CsvRecord
from report_export import (
    EXCEL_CHART_DURATION_FORMAT,
    EXCEL_DURATION_FORMAT,
    export_csv,
    export_excel,
    format_report_duration,
)


class ReportExportTests(unittest.TestCase):
    @staticmethod
    def _record(
        record_id: str,
        *,
        deleted: bool = False,
        origin: str = "TIMER",
        project: str = "Projeto Exportação",
        activity: str = "Teste",
        start: datetime | None = None,
        duration_seconds: int = 3600,
    ) -> CsvRecord:
        start = start or datetime(2026, 7, 13, 8, 0)
        return CsvRecord(
            record_id=record_id,
            user="Usuário Teste",
            origin=origin,
            project=project,
            activity_type=activity,
            description="Validar exportação",
            start=start,
            end=start + timedelta(seconds=duration_seconds),
            duration_seconds=duration_seconds,
            observation="Observação",
            computer="PC-TESTE",
            registered_at=(start + timedelta(seconds=duration_seconds)).strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
            source_file=f"dados-ficticios/{start:%Y-%m}.csv",
            deleted=deleted,
            deletion_reason="Registro incorreto" if deleted else "",
            deleted_at=(
                (start + timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S")
                if deleted
                else ""
            ),
            deleted_by="Usuário Teste" if deleted else "",
            audit_action_id="acao-1" if deleted else "",
        )

    @classmethod
    def _regression_entries(cls) -> list[tuple[str, CsvRecord]]:
        return [
            (
                "Ana",
                cls._record(
                    "assinatura-divergencia-timer",
                    project="Assinatura Eletrônica",
                    activity="Demandas de divergência",
                    start=datetime(2026, 7, 10, 8, 0),
                    duration_seconds=80 * 60,
                ),
            ),
            (
                "Ana",
                cls._record(
                    "assinatura-certificado-manual",
                    project="Assinatura Eletrônica",
                    activity="Validação de certificado",
                    origin="MANUAL",
                    start=datetime(2026, 7, 11, 9, 0),
                    duration_seconds=130 * 60,
                ),
            ),
            (
                "Ana",
                cls._record(
                    "assinatura-divergencia-excluida",
                    project="Assinatura Eletrônica",
                    activity="Demandas de divergência",
                    deleted=True,
                    start=datetime(2026, 7, 12, 10, 0),
                    duration_seconds=435 * 60,
                ),
            ),
            (
                "Bruno",
                cls._record(
                    "portal-implantacao-timer",
                    project="Portal do Cliente",
                    activity="Atendimento de implantação",
                    start=datetime(2026, 7, 10, 13, 0),
                    duration_seconds=225 * 60,
                ),
            ),
            (
                "Bruno",
                cls._record(
                    "portal-homologacao-manual",
                    project="Portal do Cliente",
                    activity="Homologação externa",
                    origin="MANUAL",
                    start=datetime(2026, 7, 13, 8, 0),
                    duration_seconds=50 * 60,
                ),
            ),
            (
                "Bruno",
                cls._record(
                    "portal-divergencia-timer",
                    project="Portal do Cliente",
                    activity="Demandas de divergência",
                    start=datetime(2026, 7, 14, 8, 0),
                    duration_seconds=310 * 60,
                ),
            ),
        ]

    @staticmethod
    def _chart_top_emu(sheet, chart) -> int:
        default_height = sheet.sheet_format.defaultRowHeight or 15
        rows_before_anchor = chart.anchor._from.row
        height_points = sum(
            sheet.row_dimensions[row].height or default_height
            for row in range(1, rows_before_anchor + 1)
        )
        return round(height_points * 12700) + chart.anchor._from.rowOff

    @classmethod
    def _export_regression_workbook(cls, target: Path) -> None:
        export_excel(
            target,
            cls._regression_entries(),
            date(2026, 7, 14),
            {
                "projeto": "Assinatura Eletrônica",
                "tipo": "Todos",
                "origem": "Todos",
                "status": "Todos",
            },
            generated_at=datetime(2026, 7, 15, 10, 0),
            period_label="10 a 14 de julho de 2026",
        )

    def test_report_duration_rounds_to_nearest_minute(self):
        self.assertEqual(format_report_duration(round(6.47 * 3600)), "06:28 h")
        self.assertEqual(format_report_duration(round(2.92 * 3600)), "02:55 h")
        self.assertEqual(format_report_duration(26 * 3600 + 5 * 60), "26:05 h")

    def test_csv_contains_filtered_records_and_audit_columns(self):
        with tempfile.TemporaryDirectory() as temporary:
            target = Path(temporary) / "relatorio.csv"
            export_csv(
                target,
                [
                    ("Usuário monitorado", self._record("r1")),
                    ("Usuário monitorado", self._record("r2", deleted=True, origin="MANUAL")),
                ],
            )
            with target.open("r", newline="", encoding="utf-8-sig") as handle:
                rows = list(csv.DictReader(handle, delimiter=";"))
            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[0]["status"], "ATIVO")
            self.assertEqual(rows[1]["status"], "EXCLUÍDO")
            self.assertEqual(rows[1]["motivo_exclusao"], "Registro incorreto")
            self.assertEqual(rows[1]["origem_registro"], "MANUAL")
            self.assertEqual(rows[0]["duracao_formatada"], "01:00 h")

    def test_excel_has_dynamic_dashboard_and_complete_records_table(self):
        with tempfile.TemporaryDirectory() as temporary:
            target = Path(temporary) / "relatorio.xlsx"
            export_excel(
                target,
                [
                    ("Usuário monitorado", self._record("r1")),
                    ("Usuário monitorado", self._record("r2", deleted=True, origin="MANUAL")),
                ],
                date(2026, 7, 13),
                {"projeto": "Todos", "tipo": "Todos", "origem": "Todos", "status": "Todos"},
                generated_at=datetime(2026, 7, 13, 12, 0),
                period_label="Janeiro a Junho de 2026",
            )
            workbook = load_workbook(target, data_only=False)
            self.assertEqual(workbook.sheetnames, ["Dashboard", "Registros"])
            self.assertEqual(len(workbook["Dashboard"]._charts), 3)
            self.assertEqual(workbook["Dashboard"]["B4"].value, "Janeiro a Junho de 2026")
            self.assertEqual(workbook["Registros"].max_row, 3)
            self.assertEqual(workbook["Registros"]["K3"].value, "EXCLUÍDO")
            self.assertEqual(workbook["Registros"]["M3"].value, "Registro incorreto")
            self.assertTrue(workbook["Registros"].column_dimensions["T"].hidden)
            self.assertTrue(workbook["Registros"].column_dimensions["X"].hidden)
            self.assertEqual(workbook["Registros"].tables["tbRegistros"].ref, "A1:X3")
            self.assertIn("SUMPRODUCT", workbook["Dashboard"]["A14"].value)
            self.assertIn("'Registros'!", workbook["Dashboard"]["A14"].value)
            self.assertEqual(workbook["Dashboard"]["A6"].value, "Todos")
            self.assertEqual(workbook["Dashboard"]["C6"].value, "Todos")
            self.assertEqual(workbook["Dashboard"]["E6"].value, "Todos")
            self.assertEqual(workbook["Dashboard"]["G6"].value, "Todos")
            self.assertEqual(workbook["Dashboard"]["E9"].value, "Todos")
            self.assertEqual(len(workbook["Dashboard"].data_validations.dataValidation), 5)
            self.assertTrue(workbook["Dashboard"].column_dimensions["J"].hidden)
            self.assertTrue(workbook["Dashboard"].column_dimensions["Y"].hidden)
            self.assertEqual(workbook.calculation.calcMode, "auto")
            self.assertTrue(workbook.calculation.fullCalcOnLoad)
            self.assertTrue(workbook.calculation.forceFullCalc)
            self.assertTrue(workbook.calculation.calcOnSave)
            self.assertFalse(workbook.calculation.calcCompleted)
            self.assertTrue(workbook.calculation.fullPrecision)
            self.assertEqual(workbook.calculation.calcId, 0)

            # Durações são armazenadas como frações de dia e exibidas em [h]:mm,
            # sem horas decimais. Isso preserva soma, filtros e gráficos.
            self.assertEqual(workbook["Registros"]["I2"].value.total_seconds(), 3600)
            self.assertEqual(workbook["Registros"]["I2"].number_format, EXCEL_DURATION_FORMAT)
            self.assertEqual(workbook["Registros"]["T2"].value.total_seconds(), 3600)
            self.assertEqual(workbook["Registros"]["U2"].value, 1)
            self.assertEqual(workbook["Registros"]["V2"].value, 0)
            self.assertEqual(workbook["Registros"]["W2"].value, 0)
            self.assertEqual(workbook["Registros"]["T3"].value.total_seconds(), 0)
            self.assertEqual(workbook["Registros"]["U3"].value, 0)
            self.assertEqual(workbook["Registros"]["V3"].value, 0)
            self.assertEqual(workbook["Registros"]["W3"].value, 1)
            self.assertEqual(workbook["Registros"]["X2"].value, datetime(2026, 7, 13))
            self.assertTrue(workbook["Registros"].column_dimensions["H"].hidden)
            self.assertEqual(workbook["Dashboard"]["A14"].number_format, EXCEL_DURATION_FORMAT)

            # Uma tabela do Excel já contém seu próprio AutoFiltro. A planilha
            # não pode declarar outro AutoFiltro sobre o mesmo intervalo, pois
            # o Excel repara e remove a tabela ao abrir o arquivo.
            with zipfile.ZipFile(target) as archive:
                self.assertIsNone(archive.testzip())
                registros_xml = archive.read("xl/worksheets/sheet2.xml").decode("utf-8")
                self.assertNotIn("<autoFilter", registros_xml)
                self.assertIn("<tableParts", registros_xml)
                table_xml = archive.read("xl/tables/table1.xml").decode("utf-8")
                self.assertIn('name="tbRegistros"', table_xml)
                self.assertIn('ref="A1:X3"', table_xml)
                self.assertIn('<autoFilter ref="A1:X3"', table_xml)

                workbook_xml = ElementTree.fromstring(archive.read("xl/workbook.xml"))
                namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
                calculation = workbook_xml.find(f"{namespace}calcPr")
                self.assertIsNotNone(calculation)
                self.assertEqual(calculation.attrib["calcId"], "0")
                self.assertEqual(calculation.attrib["calcMode"], "auto")
                self.assertEqual(calculation.attrib["fullCalcOnLoad"], "1")
                self.assertEqual(calculation.attrib["forceFullCalc"], "1")
                self.assertEqual(calculation.attrib["calcOnSave"], "1")
                self.assertEqual(calculation.attrib["calcCompleted"], "0")

    def test_excel_regression_keeps_project_and_activity_domains_separate(self):
        with tempfile.TemporaryDirectory() as temporary:
            target = Path(temporary) / "regressao-dashboard.xlsx"
            self._export_regression_workbook(target)

            workbook = load_workbook(target, data_only=False)
            dashboard = workbook["Dashboard"]
            records = workbook["Registros"]

            projects = [dashboard[f"J{row}"].value for row in range(2, 4)]
            activities = [dashboard[f"N{row}"].value for row in range(2, 6)]
            self.assertEqual(projects, ["Assinatura Eletrônica", "Portal do Cliente"])
            self.assertEqual(
                activities,
                [
                    "Atendimento de implantação",
                    "Demandas de divergência",
                    "Homologação externa",
                    "Validação de certificado",
                ],
            )
            self.assertTrue(set(projects).isdisjoint(activities))
            self.assertEqual(dashboard["C6"].value, "Assinatura Eletrônica")

            global_project_filter = (
                "'Registros'!$C$2:$C$7=IF($C$6=\"(Sem projeto)\",\"\",$C$6)"
            )
            for row in range(2, 4):
                formula = dashboard[f"K{row}"].value
                self.assertIn("SUMPRODUCT('Registros'!$T$2:$T$7", formula)
                self.assertIn(global_project_filter, formula)
                self.assertIn(f"'Registros'!$C$2:$C$7=IF($J{row}", formula)

            for row in range(2, 6):
                formula = dashboard[f"O{row}"].value
                self.assertIn("SUMPRODUCT('Registros'!$T$2:$T$7", formula)
                self.assertIn(global_project_filter, formula)
                self.assertIn(f"'Registros'!$D$2:$D$7=IF($N{row}", formula)

            # O critério de atividade continua subordinado ao projeto global.
            # Assim, a atividade de mesmo nome no Portal não entra no total da
            # Assinatura, e atividades exclusivas do Portal permanecem zeradas.
            selected_project = "Assinatura Eletrônica"
            selected_activity_totals: dict[str, int] = {}
            for row in range(2, 8):
                if records[f"C{row}"].value != selected_project:
                    continue
                duration = records[f"T{row}"].value
                seconds = round(duration.total_seconds()) if duration else 0
                activity = records[f"D{row}"].value
                selected_activity_totals[activity] = (
                    selected_activity_totals.get(activity, 0) + seconds
                )
            self.assertEqual(selected_activity_totals["Demandas de divergência"], 80 * 60)
            self.assertEqual(selected_activity_totals.get("Atendimento de implantação", 0), 0)
            self.assertEqual(selected_activity_totals.get("Homologação externa", 0), 0)

            metric_columns = {"A14": "T", "C14": "U", "E14": "V", "G14": "W"}
            for cell, source_column in metric_columns.items():
                formula = dashboard[cell].value
                self.assertIn(
                    f"SUMPRODUCT('Registros'!${source_column}$2:${source_column}$7",
                    formula,
                )
                self.assertIn(global_project_filter, formula)

            # A fonte auditável preserva projeto C, atividade D, duração ativa
            # T e data X. O registro excluído mantém auditoria, mas soma zero.
            self.assertEqual(records["C4"].value, "Assinatura Eletrônica")
            self.assertEqual(records["D4"].value, "Demandas de divergência")
            self.assertEqual(records["K4"].value, "EXCLUÍDO")
            self.assertEqual(records["T4"].value.total_seconds(), 0)
            self.assertEqual(records["W4"].value, 1)
            self.assertEqual(records["X4"].value, datetime(2026, 7, 12))
            self.assertEqual(records.tables["tbRegistros"].ref, "A1:X7")

            daily_formula = dashboard["Y2"].value
            self.assertIn("SUMPRODUCT('Registros'!$T$2:$T$7", daily_formula)
            self.assertIn(global_project_filter, daily_formula)
            self.assertIn("'Registros'!$X$2:$X$7=$X2", daily_formula)
            self.assertIn("$X2<$A$9", daily_formula)
            self.assertIn("$X2>$C$9", daily_formula)
            self.assertIn("<=0),NA()", daily_formula)

    def test_excel_charts_have_independent_areas_labels_and_date_axis(self):
        with tempfile.TemporaryDirectory() as temporary:
            target = Path(temporary) / "regressao-graficos.xlsx"
            self._export_regression_workbook(target)

            workbook = load_workbook(target, data_only=False)
            dashboard = workbook["Dashboard"]
            project_chart, activity_chart, daily_chart = dashboard._charts

            self.assertEqual(
                [chart.anchor._from.row + 1 for chart in dashboard._charts],
                [18, 37, 56],
            )
            self.assertEqual(
                [chart.anchor._from.col + 1 for chart in dashboard._charts],
                [4, 4, 1],
            )
            project_bottom = (
                self._chart_top_emu(dashboard, project_chart) + project_chart.anchor.ext.cy
            )
            activity_top = self._chart_top_emu(dashboard, activity_chart)
            activity_bottom = (
                self._chart_top_emu(dashboard, activity_chart) + activity_chart.anchor.ext.cy
            )
            daily_top = self._chart_top_emu(dashboard, daily_chart)
            self.assertLess(project_bottom, activity_top)
            self.assertLess(activity_bottom, daily_top)
            self.assertEqual(dashboard["A18"].value, "Horas por projeto")
            self.assertEqual(dashboard["A37"].value, "Horas por atividade / tarefa")
            self.assertIn("$H$73", str(dashboard.print_area))
            self.assertEqual(dashboard.page_setup.fitToWidth, 1)
            self.assertEqual(dashboard.page_setup.fitToHeight, 2)

            for chart in (project_chart, activity_chart):
                self.assertTrue(chart.dLbls.showVal)
                self.assertEqual(chart.dLbls.dLblPos, "outEnd")
                self.assertEqual(chart.dLbls.numFmt, EXCEL_CHART_DURATION_FORMAT)
                self.assertFalse(chart.dLbls.showCatName)
                self.assertFalse(chart.dLbls.showSerName)

            self.assertEqual(
                project_chart.series[0].val.numRef.f,
                "'Dashboard'!$B$20:$B$21",
            )
            self.assertEqual(
                project_chart.series[0].cat.numRef.f,
                "'Dashboard'!$A$20:$A$21",
            )
            self.assertEqual(
                activity_chart.series[0].val.numRef.f,
                "'Dashboard'!$B$39:$B$42",
            )
            self.assertEqual(
                activity_chart.series[0].cat.numRef.f,
                "'Dashboard'!$A$39:$A$42",
            )
            self.assertIn("INDEX($J$2:$J$3", dashboard["A20"].value)
            self.assertIn("INDEX($K$2:$K$3", dashboard["B20"].value)
            self.assertIn("INDEX($N$2:$N$5", dashboard["A39"].value)
            self.assertIn("INDEX($O$2:$O$5", dashboard["B39"].value)

            self.assertIsInstance(daily_chart.x_axis, DateAxis)
            self.assertEqual(daily_chart.x_axis.numFmt.formatCode, "dd/mm")
            self.assertEqual(daily_chart.x_axis.baseTimeUnit, "days")
            self.assertEqual(
                daily_chart.series[0].val.numRef.f,
                "'Dashboard'!$Y$2:$Y$6",
            )
            self.assertEqual(
                daily_chart.series[0].cat.numRef.f,
                "'Dashboard'!$X$2:$X$6",
            )
            self.assertTrue(daily_chart.dLbls.showVal)
            self.assertEqual(daily_chart.dLbls.dLblPos, "t")
            self.assertEqual(daily_chart.dLbls.numFmt, EXCEL_CHART_DURATION_FORMAT)
            self.assertEqual(daily_chart.series[0].marker.symbol, "circle")
            self.assertEqual(daily_chart.series[0].marker.size, 7)

            chart_dates = [dashboard[f"X{row}"].value for row in range(2, 7)]
            self.assertEqual(
                chart_dates,
                [
                    datetime(2026, 7, 10),
                    datetime(2026, 7, 11),
                    datetime(2026, 7, 12),
                    datetime(2026, 7, 13),
                    datetime(2026, 7, 14),
                ],
            )
            self.assertEqual(chart_dates, sorted(set(chart_dates)))

            with zipfile.ZipFile(target) as archive:
                chart_xmls = [
                    archive.read(f"xl/charts/chart{index}.xml")
                    for index in range(1, 4)
                ]
                self.assertTrue(all(b"<dLbls>" in xml for xml in chart_xmls))
                self.assertIn(b"<dateAx>", chart_xmls[2])


if __name__ == "__main__":
    unittest.main()
