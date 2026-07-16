from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Sequence

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.axis import DateAxis
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError:
    Workbook = None
    BarChart = LineChart = Reference = None
    DateAxis = DataLabelList = None
    Alignment = Border = Font = PatternFill = Side = None
    DataValidation = None
    Table = TableStyleInfo = None
    get_column_letter = None

from csv_reader import CsvRecord

APP_TITLE = "Timer Task Master"

NAVY = "13213A"
BLUE = "1D4ED8"
LIGHT_BLUE = "E8F0FF"
PALE = "F6F8FC"
WHITE = "FFFFFF"
TEXT = "172033"
MUTED = "667085"
RED = "B42318"
LIGHT_RED = "FEE4E2"
BORDER = "D7DEEA"

RecordEntry = tuple[str, CsvRecord]

EXCEL_DURATION_FORMAT = '[h]:mm "h"'
EXCEL_CHART_DURATION_FORMAT = '[h]:mm "h";-[h]:mm "h";;'
SECONDS_PER_DAY = 24 * 60 * 60

PROJECT_RANKING_ROW = 18
ACTIVITY_RANKING_ROW = 37
DAILY_CHART_ROW = 56
DASHBOARD_PRINT_END_ROW = 73


def format_report_duration(total_seconds: int) -> str:
    """Format a report duration rounded to the nearest minute."""
    total_minutes = round(max(0, int(total_seconds)) / 60)
    hours, minutes = divmod(total_minutes, 60)
    return f"{hours:02d}:{minutes:02d} h"


def _excel_duration_value(total_seconds: int) -> float:
    """Convert seconds to a real Excel duration value.

    Excel stores durations as fractions of a day. The [h]:mm number format
    displays the rounded minute while preserving the underlying seconds for
    sums, filtered totals and charts.
    """
    return max(0, int(total_seconds)) / SECONDS_PER_DAY


def _configure_excel_recalculation(workbook: Workbook) -> None:
    """Ask spreadsheet applications to rebuild formula and chart results.

    openpyxl writes formulas but does not calculate them. A zero calculation
    engine id and an incomplete calculation state make Excel discard stale
    cached results and perform a full calculation when the workbook opens.
    Optional properties are guarded to preserve compatibility across the
    supported openpyxl 3.1.x range.
    """
    calculation = workbook.calculation
    settings = {
        "calcMode": "auto",
        "fullCalcOnLoad": True,
        "forceFullCalc": True,
        "calcOnSave": True,
        "calcCompleted": False,
        "fullPrecision": True,
        "calcId": 0,
    }
    for attribute, value in settings.items():
        if hasattr(calculation, attribute):
            setattr(calculation, attribute, value)


def _normalized_path(path: str | Path, suffix: str) -> Path:
    target = Path(path)
    if target.suffix.casefold() != suffix.casefold():
        target = target.with_suffix(suffix)
    target.parent.mkdir(parents=True, exist_ok=True)
    return target


def _record_rows(entries: Iterable[RecordEntry]) -> list[list[object]]:
    rows: list[list[object]] = []
    for monitored_user, record in entries:
        rows.append(
            [
                monitored_user,
                record.user,
                record.project,
                record.activity_type,
                record.description,
                record.start,
                record.end,
                record.duration_seconds,
                format_report_duration(record.duration_seconds),
                record.origin,
                "EXCLUÍDO" if record.deleted else "ATIVO",
                record.observation,
                record.deletion_reason,
                record.deleted_by,
                record.deleted_at,
                record.computer,
                record.registered_at,
                record.record_id,
                record.source_file,
            ]
        )
    return rows


def _excel_record_rows(entries: Iterable[RecordEntry]) -> list[list[object]]:
    """Return report rows plus helper columns used by Dashboard formulas.

    The helper columns remain hidden in Excel. Numeric flags and a date-only
    value keep the management formulas simple and auditable.
    """
    rows: list[list[object]] = []
    for row in _record_rows(entries):
        deleted = row[10] == "EXCLUÍDO"
        manual = row[9] == "MANUAL"
        duration_seconds = int(row[7] or 0)
        excel_row = list(row)
        # Column I is a real Excel duration instead of a preformatted string.
        excel_row[8] = _excel_duration_value(duration_seconds)
        rows.append(
            excel_row
            + [
                0 if deleted else _excel_duration_value(duration_seconds),
                0 if deleted else 1,
                1 if manual and not deleted else 0,
                1 if deleted else 0,
                row[5].date(),
            ]
        )
    return rows


def export_csv(path: str | Path, entries: Sequence[RecordEntry]) -> Path:
    target = _normalized_path(path, ".csv")
    headers = [
        "usuario_monitorado",
        "usuario_registro",
        "projeto",
        "tipo_atividade",
        "descricao",
        "inicio",
        "fim",
        "duracao_segundos",
        "duracao_formatada",
        "origem_registro",
        "status",
        "observacao",
        "motivo_exclusao",
        "excluido_por",
        "excluido_em",
        "computador",
        "data_registro",
        "registro_id",
        "arquivo_origem",
    ]
    with target.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        for row in _record_rows(entries):
            serialized = list(row)
            serialized[5] = row[5].strftime("%Y-%m-%d %H:%M:%S")
            serialized[6] = row[6].strftime("%Y-%m-%d %H:%M:%S")
            writer.writerow(serialized)
    return target


def _style_title(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True, size=18)
    cell.alignment = Alignment(vertical="center")


def _style_header_row(sheet, row: int, start_col: int, end_col: int) -> None:
    thin = Side(style="thin", color=BORDER)
    for column in range(start_col, end_col + 1):
        cell = sheet.cell(row=row, column=column)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def _autosize(sheet, min_width: int = 11, max_width: int = 45) -> None:
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        length = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            text = value.strftime("%d/%m/%Y %H:%M") if isinstance(value, datetime) else str(value)
            length = max(length, max((len(line) for line in text.splitlines()), default=0))
        sheet.column_dimensions[letter].width = max(min_width, min(max_width, length + 2))


def _add_table(sheet, start_row: int, end_row: int, end_col: int, name: str) -> None:
    if end_row <= start_row:
        return
    reference = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _write_detail_sheet(
    workbook: Workbook,
    title: str,
    entries: Sequence[RecordEntry],
    table_name: str,
):
    sheet = workbook.create_sheet(title)
    headers = [
        "Usuário monitorado",
        "Usuário do registro",
        "Projeto",
        "Tipo de atividade",
        "Descrição",
        "Início",
        "Fim",
        "Duração (s)",
        "Duração",
        "Origem",
        "Status",
        "Observação",
        "Motivo da exclusão",
        "Excluído por",
        "Excluído em",
        "Computador",
        "Data de registro",
        "UUID",
        "Arquivo de origem",
        "Horas contabilizadas",
        "Registros contabilizados",
        "Registros manuais",
        "Registros excluídos",
        "Data do registro",
    ]
    sheet.append(headers)
    _style_header_row(sheet, 1, 1, len(headers))

    excel_rows = _excel_record_rows(entries)
    for row_index, values in enumerate(excel_rows, start=2):
        sheet.append(values)
        sheet.cell(row_index, 6).number_format = "dd/mm/yyyy hh:mm"
        sheet.cell(row_index, 7).number_format = "dd/mm/yyyy hh:mm"
        sheet.cell(row_index, 9).number_format = EXCEL_DURATION_FORMAT
        sheet.cell(row_index, 20).number_format = EXCEL_DURATION_FORMAT
        sheet.cell(row_index, 24).number_format = "dd/mm/yyyy"
        if values[10] == "EXCLUÍDO":
            for column in range(1, 20):
                cell = sheet.cell(row_index, column)
                cell.fill = PatternFill("solid", fgColor=LIGHT_RED)
                cell.font = Font(color=RED)

    if not entries:
        sheet.cell(2, 1, "Nenhum registro para os filtros selecionados.")
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=19)
        sheet.cell(2, 1).font = Font(color=MUTED, italic=True)

    sheet.freeze_panes = "A2"
    # A tabela já fornece os filtros de coluna. Não configure também
    # worksheet.auto_filter no mesmo intervalo.
    _add_table(sheet, 1, len(entries) + 1, len(headers), table_name)
    _autosize(sheet)
    sheet.column_dimensions["E"].width = 34
    sheet.column_dimensions["L"].width = 34
    sheet.column_dimensions["M"].width = 30
    sheet.column_dimensions["S"].width = 48

    # Helper columns power the dynamic Dashboard but do not clutter the
    # exported report. They remain part of the table so every calculation has
    # an auditable source inside Registros.
    # Raw seconds remain available in the file for auditing, but the visible
    # duration is the real Excel time value in column I.
    sheet.column_dimensions["H"].hidden = True
    for column in ("T", "U", "V", "W", "X"):
        sheet.column_dimensions[column].hidden = True
    return sheet


def _dashboard_metric_formula(
    end_row: int,
    value_column: str,
    *,
    extra_column: str | None = None,
    extra_cell: str | None = None,
    blank_label: str | None = None,
) -> str:
    """Build a formula driven by the editable Dashboard filter cells."""

    def data_range(column: str) -> str:
        return f"'Registros'!${column}$2:${column}${end_row}"

    user_value = 'IF($A$6="(Sem usuário)","",$A$6)'
    project_value = 'IF($C$6="(Sem projeto)","",$C$6)'
    activity_value = 'IF($E$6="(Sem atividade)","",$E$6)'
    origin_value = 'IF($G$6="(Sem origem)","",$G$6)'
    factors = [
        f'IF($A$6="Todos",1,--({data_range("A")}={user_value}))',
        f'IF($C$6="Todos",1,--({data_range("C")}={project_value}))',
        f'IF($E$6="Todos",1,--({data_range("D")}={activity_value}))',
        f'IF($G$6="Todos",1,--({data_range("J")}={origin_value}))',
        (
            f'IF($E$9="Todos",1,--({data_range("K")}='
            'IF($E$9="Ativos","ATIVO","EXCLUÍDO")))'
        ),
        f'--({data_range("X")}>=$A$9)',
        f'--({data_range("X")}<=$C$9)',
    ]
    if extra_column and extra_cell:
        criterion = (
            f'IF({extra_cell}="{blank_label}","",{extra_cell})'
            if blank_label
            else extra_cell
        )
        factors.append(f'--({data_range(extra_column)}={criterion})')

    return f'=SUMPRODUCT({data_range(value_column)}*{"*".join(factors)})'


def _filter_value(value: str | None, options: Sequence[str]) -> str:
    return value if value in options else "Todos"


def _add_dashboard_validation(
    sheet,
    target_cell: str,
    source_column: str,
    values: Sequence[str],
) -> None:
    for row, value in enumerate(values, start=2):
        sheet.cell(row=row, column=ord(source_column) - ord("A") + 1, value=value)
    end_row = len(values) + 1
    validation = DataValidation(
        type="list",
        formula1=f"'Dashboard'!${source_column}$2:${source_column}${end_row}",
        allow_blank=False,
    )
    validation.errorTitle = "Filtro inválido"
    validation.error = "Escolha um valor da lista."
    validation.promptTitle = "Filtro do Dashboard"
    validation.prompt = "Selecione um valor para recalcular os indicadores e gráficos."
    validation.showErrorMessage = True
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    validation.add(sheet[target_cell])


def _style_dashboard_filter(sheet, label_range: str, input_range: str, label: str, value) -> None:
    label_cell = label_range.split(":", 1)[0]
    input_cell = input_range.split(":", 1)[0]
    sheet.merge_cells(label_range)
    sheet.merge_cells(input_range)
    sheet[label_cell] = label
    sheet[input_cell] = value
    sheet[label_cell].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    sheet[label_cell].font = Font(bold=True, color=NAVY)
    sheet[label_cell].alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color=BLUE)
    for row in sheet[input_range]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=WHITE)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    sheet[input_cell].font = Font(bold=True, color=TEXT)
    sheet[input_cell].alignment = Alignment(horizontal="center", vertical="center")


def _write_dashboard_sheet(
    sheet,
    entries: Sequence[RecordEntry],
    report_date: date,
    filters: dict[str, str],
    generated_at: datetime,
    period_label: str | None,
) -> None:
    records = [record for _user, record in entries]
    user_values = sorted({user or "(Sem usuário)" for user, _record in entries}, key=str.casefold)
    project_values = sorted(
        {record.project or "(Sem projeto)" for record in records},
        key=str.casefold,
    )
    activity_values = sorted(
        {record.activity_type or "(Sem atividade)" for record in records},
        key=str.casefold,
    )
    origin_values = sorted(
        {record.origin or "(Sem origem)" for record in records},
        key=str.casefold,
    )
    record_dates = sorted({record.start.date() for record in records})
    start_date = record_dates[0] if record_dates else report_date
    end_date = record_dates[-1] if record_dates else report_date

    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 85
    sheet.merge_cells("A1:H2")
    sheet["A1"] = f"{APP_TITLE} — Dashboard de horas"
    _style_title(sheet["A1"])
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 12

    sheet["A4"] = "Período exportado"
    sheet["A4"].font = Font(bold=True, color=NAVY)
    sheet.merge_cells("B4:D4")
    sheet["B4"] = period_label or report_date
    if not period_label:
        sheet["B4"].number_format = "dd/mm/yyyy"
    sheet["E4"] = "Gerado em"
    sheet["E4"].font = Font(bold=True, color=NAVY)
    sheet.merge_cells("F4:H4")
    sheet["F4"] = generated_at
    sheet["F4"].number_format = "dd/mm/yyyy hh:mm"

    user_options = ["Todos", *user_values]
    project_options = ["Todos", *project_values]
    activity_options = ["Todos", *activity_values]
    origin_options = ["Todos", *origin_values]
    status_options = ["Todos", "Ativos", "Excluídos"]

    _style_dashboard_filter(sheet, "A5:B5", "A6:B6", "Usuário", "Todos")
    _style_dashboard_filter(
        sheet,
        "C5:D5",
        "C6:D6",
        "Projeto",
        _filter_value(filters.get("projeto"), project_options),
    )
    _style_dashboard_filter(
        sheet,
        "E5:F5",
        "E6:F6",
        "Atividade / tarefa",
        _filter_value(filters.get("tipo"), activity_options),
    )
    _style_dashboard_filter(
        sheet,
        "G5:H5",
        "G6:H6",
        "Origem",
        _filter_value(filters.get("origem"), origin_options),
    )
    _style_dashboard_filter(sheet, "A8:B8", "A9:B9", "Data inicial", start_date)
    _style_dashboard_filter(sheet, "C8:D8", "C9:D9", "Data final", end_date)
    _style_dashboard_filter(
        sheet,
        "E8:F8",
        "E9:F9",
        "Status",
        _filter_value(filters.get("status"), status_options),
    )
    _style_dashboard_filter(sheet, "G8:H8", "G9:H9", "Fonte dos dados", "Registros")
    sheet["A9"].number_format = "dd/mm/yyyy"
    sheet["C9"].number_format = "dd/mm/yyyy"
    sheet["G9"].font = Font(bold=True, color=BLUE)

    _add_dashboard_validation(sheet, "A6", "R", user_options)
    _add_dashboard_validation(sheet, "C6", "S", project_options)
    _add_dashboard_validation(sheet, "E6", "T", activity_options)
    _add_dashboard_validation(sheet, "G6", "U", origin_options)
    _add_dashboard_validation(sheet, "E9", "V", status_options)

    sheet.merge_cells("A11:H11")
    sheet["A11"] = (
        "Altere os filtros acima: indicadores e gráficos recalculam usando a tabela completa da aba Registros."
    )
    sheet["A11"].font = Font(color=MUTED, italic=True, size=10)
    sheet["A11"].alignment = Alignment(horizontal="center", vertical="center")

    card_specs = [
        (1, "Horas contabilizadas", "T", EXCEL_DURATION_FORMAT),
        (3, "Registros ativos", "U", "0"),
        (5, "Registros manuais", "V", "0"),
        (7, "Registros excluídos", "W", "0"),
    ]
    end_row = max(2, len(entries) + 1)
    border_side = Side(style="thin", color=BORDER)
    for column, label, metric_column, number_format in card_specs:
        sheet.merge_cells(
            start_row=13,
            start_column=column,
            end_row=13,
            end_column=column + 1,
        )
        sheet.merge_cells(
            start_row=14,
            start_column=column,
            end_row=15,
            end_column=column + 1,
        )
        label_cell = sheet.cell(row=13, column=column, value=label)
        value_cell = sheet.cell(
            row=14,
            column=column,
            value=_dashboard_metric_formula(end_row, metric_column),
        )
        label_cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        label_cell.font = Font(bold=True, color=NAVY)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.fill = PatternFill("solid", fgColor=WHITE)
        value_cell.font = Font(bold=True, color=TEXT, size=18)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = number_format
        for row in sheet.iter_rows(min_row=13, max_row=15, min_col=column, max_col=column + 1):
            for cell in row:
                cell.border = Border(
                    top=border_side,
                    bottom=border_side,
                    left=border_side,
                    right=border_side,
                )

    # Formula-backed helper lists feed the visible rankings and charts. They
    # stay on Dashboard so the workbook has only the two user-facing sheets.
    sheet["J1"] = "Projeto"
    sheet["K1"] = "Horas"
    sheet["L1"] = "Ordem"
    for row, project in enumerate(project_values, start=2):
        sheet.cell(row=row, column=10, value=project)
        sheet.cell(
            row=row,
            column=11,
            value=_dashboard_metric_formula(
                end_row,
                "T",
                extra_column="C",
                extra_cell=f"$J{row}",
                blank_label="(Sem projeto)",
            ),
        ).number_format = EXCEL_DURATION_FORMAT
        sheet.cell(row=row, column=12, value=f'=IF(K{row}>0,K{row}+ROW()/1000000000,0)')

    sheet["N1"] = "Atividade"
    sheet["O1"] = "Horas"
    sheet["P1"] = "Ordem"
    for row, activity in enumerate(activity_values, start=2):
        sheet.cell(row=row, column=14, value=activity)
        sheet.cell(
            row=row,
            column=15,
            value=_dashboard_metric_formula(
                end_row,
                "T",
                extra_column="D",
                extra_cell=f"$N{row}",
                blank_label="(Sem atividade)",
            ),
        ).number_format = EXCEL_DURATION_FORMAT
        sheet.cell(row=row, column=16, value=f'=IF(O{row}>0,O{row}+ROW()/1000000000,0)')

    def add_ranking(
        title: str,
        label: str,
        source_label_column: str,
        source_value_column: str,
        source_score_column: str,
        source_count: int,
        title_row: int,
        chart_anchor: str,
    ) -> None:
        sheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=2)
        sheet.cell(row=title_row, column=1, value=title)
        sheet.cell(row=title_row, column=1).font = Font(bold=True, color=NAVY, size=12)
        header_row = title_row + 1
        sheet.cell(row=header_row, column=1, value=label)
        sheet.cell(row=header_row, column=2, value="Horas")
        _style_header_row(sheet, header_row, 1, 2)

        if not source_count:
            sheet.merge_cells(
                start_row=header_row + 1,
                start_column=1,
                end_row=header_row + 2,
                end_column=2,
            )
            sheet.cell(row=header_row + 1, column=1, value="Sem registros para o período.")
            sheet.cell(row=header_row + 1, column=1).font = Font(color=MUTED, italic=True)
            return

        source_end = source_count + 1
        ranking_count = min(10, source_count)
        first_data_row = header_row + 1
        last_data_row = first_data_row + ranking_count - 1
        for row in range(first_data_row, last_data_row + 1):
            rank = f"ROWS($A${first_data_row}:A{row})"
            score_range = f"${source_score_column}$2:${source_score_column}${source_end}"
            label_range = f"${source_label_column}$2:${source_label_column}${source_end}"
            score = f"LARGE({score_range},{rank})"
            sheet.cell(
                row=row,
                column=1,
                value=(
                    f'=IFERROR(IF({score}<=0,"",INDEX({label_range},'
                    f'MATCH({score},{score_range},0))),"")'
                ),
            )
            sheet.cell(
                row=row,
                column=2,
                value=(
                    f'=IF($A{row}="",NA(),INDEX(${source_value_column}$2:'
                    f'${source_value_column}${source_end},MATCH($A{row},{label_range},0)))'
                ),
            ).number_format = EXCEL_DURATION_FORMAT
            if (row - first_data_row) % 2:
                for column in (1, 2):
                    sheet.cell(row=row, column=column).fill = PatternFill("solid", fgColor=PALE)

        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = title
        chart.x_axis.title = "Duração"
        chart.x_axis.numFmt = EXCEL_DURATION_FORMAT
        chart.legend = None
        chart.height = 7.6
        chart.width = 12.5
        chart.visible_cells_only = True
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.dLblPos = "outEnd"
        chart.dLbls.numFmt = EXCEL_CHART_DURATION_FORMAT
        chart.dLbls.showLegendKey = False
        chart.dLbls.showCatName = False
        chart.dLbls.showSerName = False
        chart.dLbls.showPercent = False
        chart.dLbls.showBubbleSize = False
        chart.dLbls.showLeaderLines = False
        chart.add_data(
            Reference(sheet, min_col=2, min_row=header_row, max_row=last_data_row),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(sheet, min_col=1, min_row=first_data_row, max_row=last_data_row)
        )
        sheet.add_chart(chart, chart_anchor)

    add_ranking(
        "Horas por projeto",
        "Projeto",
        "J",
        "K",
        "L",
        len(project_values),
        PROJECT_RANKING_ROW,
        f"D{PROJECT_RANKING_ROW}",
    )
    add_ranking(
        "Horas por atividade / tarefa",
        "Atividade / tarefa",
        "N",
        "O",
        "P",
        len(activity_values),
        ACTIVITY_RANKING_ROW,
        f"D{ACTIVITY_RANKING_ROW}",
    )

    sheet["X1"] = "Data"
    sheet["Y1"] = "Horas"
    for row, record_day in enumerate(record_dates, start=2):
        sheet.cell(row=row, column=24, value=record_day).number_format = "dd/mm/yyyy"
        metric_formula = _dashboard_metric_formula(
            end_row,
            "T",
            extra_column="X",
            extra_cell=f"$X{row}",
        )
        metric_expression = metric_formula[1:]
        sheet.cell(
            row=row,
            column=25,
            value=(
                f'=IF(OR($X{row}<$A$9,$X{row}>$C$9,'
                f'{metric_expression}<=0),NA(),{metric_expression})'
            ),
        ).number_format = EXCEL_DURATION_FORMAT

    if record_dates:
        trend = LineChart()
        trend.style = 10
        trend.title = "Horas por dia"
        trend.y_axis.title = "Duração"
        trend.y_axis.numFmt = EXCEL_DURATION_FORMAT
        trend.x_axis = DateAxis(axId=10, axPos="b", crossAx=100)
        trend.x_axis.title = "Data"
        trend.x_axis.number_format = "dd/mm"
        trend.x_axis.baseTimeUnit = "days"
        trend.legend = None
        trend.height = 8.0
        trend.width = 19.5
        trend.visible_cells_only = False
        trend.display_blanks = "gap"
        trend.dLbls = DataLabelList()
        trend.dLbls.showVal = True
        trend.dLbls.dLblPos = "t"
        trend.dLbls.numFmt = EXCEL_CHART_DURATION_FORMAT
        trend.dLbls.showLegendKey = False
        trend.dLbls.showCatName = False
        trend.dLbls.showSerName = False
        trend.dLbls.showPercent = False
        trend.dLbls.showBubbleSize = False
        trend.dLbls.showLeaderLines = False
        trend.add_data(
            Reference(sheet, min_col=25, min_row=1, max_row=len(record_dates) + 1),
            titles_from_data=True,
        )
        trend.set_categories(
            Reference(sheet, min_col=24, min_row=2, max_row=len(record_dates) + 1)
        )
        if trend.series:
            trend.series[0].marker.symbol = "circle"
            trend.series[0].marker.size = 7
            trend.series[0].graphicalProperties.line.solidFill = BLUE
        sheet.add_chart(trend, f"A{DAILY_CHART_ROW}")

    for column in range(1, 9):
        sheet.column_dimensions[get_column_letter(column)].width = 15
    for column in range(10, 26):
        sheet.column_dimensions[get_column_letter(column)].hidden = True
    sheet.row_dimensions[6].height = 25
    sheet.row_dimensions[9].height = 25
    sheet.row_dimensions[11].height = 24
    sheet.row_dimensions[14].height = 24
    sheet.row_dimensions[15].height = 24
    sheet.freeze_panes = "A4"
    sheet.print_area = f"A1:H{DASHBOARD_PRINT_END_ROW}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 2
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def export_excel(
    path: str | Path,
    entries: Sequence[RecordEntry],
    report_date: date,
    filters: dict[str, str],
    generated_at: datetime | None = None,
    period_label: str | None = None,
) -> Path:
    if Workbook is None:
        raise RuntimeError(
            "A dependência openpyxl não está instalada. Feche o aplicativo e execute "
            "run_timertaskmaster.bat ou run_debug.bat para instalar as dependências."
        )

    target = _normalized_path(path, ".xlsx")
    generated_at = generated_at or datetime.now()

    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    _write_detail_sheet(workbook, "Registros", entries, "tbRegistros")
    _write_dashboard_sheet(
        dashboard,
        entries,
        report_date,
        filters,
        generated_at,
        period_label,
    )

    _configure_excel_recalculation(workbook)
    workbook.active = 0
    workbook.save(target)
    return target
